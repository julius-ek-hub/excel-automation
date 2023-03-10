import openpyxl as ex, os
from openpyxl.utils.cell import get_column_letter
from utils import to_excel, _input_, cprint, beep, same_week, resource_path, from_json_file

column_names = from_json_file('./assets/cols.json')

class Scanner:
    def __init__(self, ss_path, workbook_ms, scan_date, entity, vulnerability_param, ms_target_sheet, ss_target_sheet, scan_index):
        self.ss_path = ss_path
        self.scan_index = str(scan_index + 1)
        self.workbook_ms = workbook_ms
        self.ms_col_ids = {}
        self.ss_col_ids = {}
        self.scan_date =scan_date 
        self.ms_target_sheet = ms_target_sheet
        self.ss_target_sheet = ss_target_sheet
        self.entity = entity 
        self.vulnerability_param = vulnerability_param
        self.ms_col_names = column_names.copy()
        self.ss_col_names = column_names.copy()
        self.total_updates = {"New": 0, "Newly Carried Forward": 0, "Closed": 0}
        self.ms_existing_vulnerability_rows = []
        self.target_entities_by_ips = {}
        self.ips_target_sheet = None
        self.ips_path = None
        self.play_sound = False
        self.internal = vulnerability_param == 'Internal'

    @staticmethod
    def trim(value):
        return str(value).strip()
    
    @staticmethod
    def set(cell, value):
        cell.value = value

    @staticmethod
    def severity(value):
        if str(value).isnumeric():
            return ['None', 'Low', 'Medium', 'High', 'Critical'][int(value)]
        return value
    
    def host_key(self):
        _key = 'Host'
        if not self.internal:
            _key = 'IP'
        return _key
    
    @staticmethod
    def get_column_data(sheet, column):
        if not column:
            return tuple([])
        all = list(sheet[column])
        all.pop(0)
        return tuple(all)

    def get_column(self, sheet, possible_names):
        for col in sheet.iter_cols(1, sheet.max_column):
            value = self.trim(col[0].value).lower()
            if value and any(name.lower() == value for name in possible_names):
                return get_column_letter(col[0].column)
            
    def identify_entities_by_ips(self):

        unknown_entity_ips = []
        
        for ss_row in self.get_column_data(sheet=self.ss, column=self.ss_col_ids["Plugin"]):

            ss_row_str = str(ss_row.row)
            ss_ip_value = self.trim(self.ss[self.ss_col_ids[self.host_key()] + ss_row_str].value).lower()

            if self.target_entities_by_ips.get(ss_ip_value) or ss_ip_value in unknown_entity_ips: continue

            entity_unknon = True

            ips_path = self.ips_path or './assets/ips.xlsx'
            
            ips_ws = ex.load_workbook(resource_path(ips_path))
            ips_sheet = ips_ws.active

            if self.ips_target_sheet:
                try:
                    ips_sheet = ips_ws[self.ips_target_sheet]
                except Exception as e: raise Exception(str(e) + ' (All IPs/Entities file)')

            for ip_row in self.get_column_data(sheet=ips_sheet, column='A'):

                ip_row_str = str(ip_row.row)

                ip = self.trim(ips_sheet['A' + ip_row_str].value)
                entity = self.trim(ips_sheet['B' + ip_row_str].value).upper().replace('_', ' ')
                entity = 'KATIM' if entity in ['D14', 'DIGITAL14', 'D14/KATIM'] else ('EDGE' if entity == 'EDGE CORP' else ('KNOWLEDGE POINT' if entity == 'KPOINT' else entity))

                if ip == ss_ip_value:
                    self.target_entities_by_ips[ss_ip_value] = entity
                    entity_unknon = False
                    break

            if entity_unknon:

                for ms_row in self.get_column_data(sheet=self.ms, column=self.ms_col_ids["Host"]):

                    ms_row_str = str(ms_row.row)
                    ms_ip_value = self.trim(self.ms[self.ms_col_ids["Host"] + ms_row_str].value).lower()
                    ms_entity_value = self.trim(self.ms[self.ms_col_ids["Entity"] + ms_row_str].value)

                    if ms_ip_value == ss_ip_value and ms_entity_value:
                        self.target_entities_by_ips[ss_ip_value] = ms_entity_value
                        entity_unknon = False
                        break

            if entity_unknon:
                unknown_entity_ips.append(ss_ip_value)
                allowed_entities = from_json_file('./assets/entities.json').values()
                domain = ss_ip_value.replace('-', ' ').upper()
                try_entity = ''

                if domain.endswith('.EDGEGROUP.AE'):
                    d = domain.split('.EDGEGROUP.AE')
                    dv = d[0].strip()
                    dv = 'KNOWLEDGE POINT' if dv == 'KP' else dv
                    try_entity = dv if dv in allowed_entities else ''
                elif domain.endswith('.AE'):
                   d = domain.split('.AE')
                   dv = d[0].strip()
                   dv = 'KNOWLEDGE POINT' if dv == 'KP' else ('EDGE' if dv == 'EDGEGROUP' else ('BEACON RED' if dv == 'BEACONRED' else dv))
                   try_entity = dv if dv in allowed_entities else ''

                self.target_entities_by_ips[ss_ip_value] = try_entity

        if len(unknown_entity_ips) > 0 :
            cprint('Could not find Entity in both External IPs sheet and mastersheet for the following IPs - ' + str(unknown_entity_ips) + ', their values will be blanc.' + (' However, some of the IPs are in the form of domain names. In such cases, the program tries to extract the entity name from the domain name.' if any((ip.endswith('.ae') or ip.endswith('.com')) for ip in unknown_entity_ips) else ''), 'warning')

        self.target_entities_by_ips['entities'] = list(dict.fromkeys(list(self.target_entities_by_ips.values())))


    def get_column_by_all_means(self, key: str, sheet: str, col_names, sheet_name: str = 'Mastersheet', important: bool = True):
        col = self.get_column(sheet, col_names[key])
        label = col_names[key][0]
        
        if not col and important:
            beep(self.play_sound)
            from_user = _input_(sheet_name + ' ' + label + ' column doesn\'t exist for value ' + ' or '.join(col_names[key]) + ', check ' + sheet_name + ' and enter the title for ' + label + ' column: ')
            if from_user:
                col_names[key].append(from_user)
            return self.get_column_by_all_means(key, sheet, col_names, sheet_name, important)
        return col

    def get_columns(self):

        if len(self.ms_col_ids) == 0:
            cprint('Identifying columns.... (Mastersheet)')
            for key in self.ms_col_names.keys():
                important = True
                if key in ['PN','NBN', 'Description', 'Solution', 'DD', 'CVE', 'IP']: important = False
                self.ms_col_ids[key] = self.get_column_by_all_means(sheet=self.ms, key=key, col_names=self.ms_col_names, important=important)
                

        cprint('Identifying columns.... (Scansheet ' + self.scan_index + ')')

        for key in self.ss_col_names.keys():
            important = False
            if key in [self.host_key(), 'Plugin', 'Severity']: important = True
            self.ss_col_ids[key] = self.get_column_by_all_means(sheet=self.ss, key=key, col_names=self.ss_col_names, important=important, sheet_name='Scansheet ' + self.scan_index)


    def check_mastersheet_with_scansheet(self):

        cf = 'Carried Forward'
        pcd = 'Patched'

        for ms_row in self.get_column_data(sheet=self.ms, column=self.ms_col_ids["Plugin"]):

            ms_row_str = str(ms_row.row)

            ms_cd_cell = self.ms[self.ms_col_ids["CD"] + ms_row_str]
            ms_status_cell = self.ms[self.ms_col_ids["Status"] + ms_row_str]

            ms_plugin_value = self.trim(ms_row.value).lower()
            ms_host_value = self.trim(self.ms[self.ms_col_ids["Host"] + ms_row_str].value).lower()
            ms_severity_value = self.severity(self.trim(self.ms[self.ms_col_ids["Severity"] + ms_row_str].value)).lower()
            ms_status_value = self.trim(ms_status_cell.value).lower()

            if ms_severity_value.lower() == 'none' or not (ms_plugin_value or ms_host_value or ms_severity_value): continue

            closed = ms_cd_cell.value or ms_status_value == pcd.lower()

            target_vp = self.vulnerability_param == self.trim(self.ms[self.ms_col_ids["VP"] + ms_row_str].value)
            target_entity = self.trim(self.ms[self.ms_col_ids["Entity"] + ms_row_str].value) in self.target_entities_by_ips['entities']
            date_created = self.trim(self.ms[self.ms_col_ids["Date"] + ms_row_str].value)

            if not (target_entity and target_vp) or closed: continue

            self.ms_existing_vulnerability_rows.append(ms_row_str)

            if same_week(date_created, self.scan_date): continue

            vulnerability_match = False

            for ss_row in self.get_column_data(sheet=self.ss, column=self.ss_col_ids["Plugin"]):

                if vulnerability_match: continue

                ss_row_str = str(ss_row.row)

                ss_plugin_value = self.trim(ss_row.value).lower()
                ss_host_value = self.trim(self.ss[self.ss_col_ids[self.host_key()] + ss_row_str].value).lower()
                ss_severity_value = self.severity(self.trim(self.ss[self.ss_col_ids["Severity"] + ss_row_str].value)).lower()

                if not (ss_plugin_value or ss_host_value or ss_severity_value): continue

                same_plugin = ms_plugin_value == ss_plugin_value
                same_host = ms_host_value == ss_host_value
                same_severity = ms_severity_value == ss_severity_value

                vulnerability_match = same_host and same_plugin and same_severity

                if (vulnerability_match):

                    ms_ncf_cell = self.ms[self.ms_col_ids["NCF"] + ms_row_str]

                    carried_forward = self.trim(ms_ncf_cell.value).lower() == cf.lower()

                    if not carried_forward:
                        self.set(ms_ncf_cell, cf)
                        self.total_updates["Carried Forward"] += 1
                    break
                    

            if not vulnerability_match:
                self.set(ms_cd_cell, self.scan_date)
                self.total_updates["Closed"] += 1

                if pcd != self.trim(ms_status_cell.value).lower():
                    self.set(ms_status_cell, pcd)


    def check_scansheet_with_mastersheet(self):

        entity = self.target_entities_by_ips['entities'][0]
        
        for ss_row in self.get_column_data(sheet=self.ss, column=self.ss_col_ids["Plugin"]):

            ss_row_str = str(ss_row.row)
            ss_plugin_value = self.trim(ss_row.value).lower()
            ss_host_value = self.trim(self.ss[self.ss_col_ids[self.host_key()] + ss_row_str].value).lower()
            ss_severity_value = self.severity(self.trim(self.ss[self.ss_col_ids["Severity"] + ss_row_str].value))

            if ss_severity_value.lower() == 'none' or not (ss_plugin_value or ss_host_value or ss_severity_value): continue

            vulnerabily_exists = False

            for ms_row in self.ms_existing_vulnerability_rows:
                ms_plugin_value = self.trim(self.ms[self.ms_col_ids["Plugin"] + ms_row].value).lower()
                ms_host_value = self.trim(self.ms[self.ms_col_ids["Host"] + ms_row].value).lower()

                same_plugin = ms_plugin_value == ss_plugin_value 
                same_host = ms_host_value == ss_host_value
                same_severity = self.severity(self.trim(self.ms[self.ms_col_ids["Severity"] + ms_row].value)).lower() == ss_severity_value.lower()

                if same_plugin and same_host and same_severity:
                    vulnerabily_exists = True
                    break

            if not vulnerabily_exists:

                if not self.internal:
                    entity = str(self.target_entities_by_ips[ss_host_value]).upper()

                ms_last_empty_row = str(self.ms.max_row + 1)

                new = {
                    'VP': self.vulnerability_param,
                    'Status': 'pending',
                    'Date': self.scan_date,
                    'Entity': entity, 
                    'NCF': 'New', 
                    'Plugin': int(ss_plugin_value), 
                    'Severity': int(ss_severity_value) if ss_severity_value.isnumeric() else ss_severity_value,
                    'Host': ss_host_value
                }

                for key, value in new.items():
                    self.set(self.ms[self.ms_col_ids[key] + ms_last_empty_row], value)

                for k in self.ms_col_names.keys():
                    ms_column = self.ms_col_ids[k]
                    ss_column = self.ss_col_ids[k]

                    if not (ss_column and ms_column) or (k in new.keys()): continue
                    self.set(self.ms[ms_column + ms_last_empty_row], self.ss[ss_column + ss_row_str].value)

                self.total_updates["New"] += 1

    def scan(self):

        ss_path = to_excel(path=self.ss_path)

        cprint('Loading Scansheet ' + self.scan_index + '.....')
        workbook_ss = ex.load_workbook(ss_path)
        cprint('Done!', 'success')

        self.ss = workbook_ss.active
        self.ms = self.workbook_ms.active

        if self.ms_target_sheet:
            try:
                self.ms = self.workbook_ms[self.ms_target_sheet]
            except Exception as e: raise Exception(str(e) + ' (Mastersheet)')
        if self.ss_target_sheet:
            try:
                self.ss = workbook_ss[self.ss_target_sheet]
            except Exception as e: raise Exception(str(e) + ' (Scansheet ' + self.scan_index + ')')

        # Identify columns
        self.get_columns()

        # Change D14 name to KATIM and rename BR
        for row in self.get_column_data(self.ms, self.ms_col_ids["Entity"]):
            entity = self.trim(row.value).upper()
            if entity in ['D14', 'DIGITAL14', 'D14/KATIM']: self.set(row, 'KATIM')
            if entity == 'BEACONRED': self.set(row, 'BEACON RED')

        cprint('Done!', 'success')

        if not self.internal:
            cprint('Identifying Entities - (scansheet ' + self.scan_index + ' with all-ips sheet).....')
            self.identify_entities_by_ips()
            cprint('Done!', 'success')
        else:
            self.target_entities_by_ips['entities'] = [self.entity]

        # # Check mastersheet with scansheet
        cprint('Scanning & updating Mastersheet with scansheet ' + self.scan_index + '.....')
        self.check_mastersheet_with_scansheet()
        cprint('Done!', 'success')

        # Check scansheet with mastersheet for new vulnerabilities
        cprint('Scanning for new vulnerabilities with scansheet ' + self.scan_index + '.....')
        self.check_scansheet_with_mastersheet()
        cprint('Done!', 'success')

    def save(self, path: str):

        cprint('Saving to ' + path + ' ....')

        if(os.path.exists(path)):
            os.unlink(path)
        self.workbook_ms.save(path)

        # Sometimes, an additional invalid file with no extension is created. So...
        try:
            os.unlink(path.replace('.xlsx', ''))
        except:
            pass
        cprint('Done!', 'success')