import openpyxl as ex, os
from openpyxl.styles import Alignment
from utils import to_excel, column_names, _input_, cprint, beep, same_week

class Scanner:
    def __init__(self, ss_path, workbook_ms, scan_date, entity, vulnerability_param, ms_target_sheet, ss_target_sheet, scan_index):
        self.ss_path = ss_path
        self.scan_index = str(scan_index + 1)
        self.workbook_ms = workbook_ms
        self.ms_col_ids = {}
        self.ss_col_ids = {}
        self.scan_date =scan_date 
        self.ms_target_sheet = ms_target_sheet
        self.ss_target_sheet = ss_target_sheet
        self.entity = entity 
        self.vulnerability_param = vulnerability_param
        self.ms_default_cols = column_names.copy()
        self.ss_default_cols = column_names.copy()
        self.non_mandatory_columns = {"ss": {}, "ms": {}}
        self.nm_column_keys = ['Plugin', 'Host', 'PN', 'Severity','NBN', 'Description', 'Solution', 'DD', 'CVE']
        self.total_updates = {"New": 0, "Newly Carried Forward": 0, "Closed": 0}
        self.ms_existing_vulnerability_rows = []
        self.play_sound = False

    @staticmethod
    def trim(value):
        return str(value).strip()
    
    @staticmethod
    def set(cell, value):
        cell.value = value
        cell.alignment = Alignment(horizontal='right')
    
    @staticmethod
    def get_column_data(sheet, column):
        if not column:
            return tuple([])
        all = list(sheet[column])
        all.pop(0)
        return tuple(all)

    def get_column(self, sheet, search_value):
        for col in sheet.iter_cols(1, sheet.max_column):
            value = self.trim(col[0].value).lower()
            if value and any(name.lower() == value for name in search_value.split('~&~&~')):
                return chr(64 + col[0].column)


    def get_column_by_all_means(self, label: str, key: str, sheet: str, default_cols, sheet_name: str = 'Mastersheet', important: bool = True):
        col = self.get_column(sheet, default_cols[key])
        
        if not col and important:
            beep(self.play_sound)
            from_user = _input_(sheet_name + ' ' + label + ' column doesn\'t exist for value ' + ' or '.join(default_cols[key].split('~&~&~')) + ', check ' + sheet_name + ' and enter the title for ' + label + ' column: ')
            if from_user:
                default_cols[key] = from_user
            return self.get_column_by_all_means(label, key, sheet, default_cols, sheet_name)
        return col

    def get_columns(self):

        # Mandatory columns

        if len(self.ms_col_ids) == 0:
            cprint('Identifying columns.... (Mastersheet)')
            self.ms_col_ids["Host"] = self.get_column_by_all_means(sheet=self.ms, key='Host', default_cols=self.ms_default_cols, label='Host')
            self.ms_col_ids["Plugin"] = self.get_column_by_all_means(sheet=self.ms, key='Plugin', default_cols=self.ms_default_cols, label='Plugin')
            self.ms_col_ids["Date"] = self.get_column_by_all_means(sheet=self.ms, key='Date', default_cols=self.ms_default_cols, label='Date')
            self.ms_col_ids["Status"] = self.get_column_by_all_means(sheet=self.ms, key='Status', default_cols=self.ms_default_cols, label='Status')
            self.ms_col_ids["NCF"] = self.get_column_by_all_means(sheet=self.ms, key='NCF', default_cols=self.ms_default_cols, label='New/Carried forward')
            self.ms_col_ids["VP"] = self.get_column_by_all_means(sheet=self.ms, key='VP', default_cols=self.ms_default_cols, label='Vulnerability parameter')
            self.ms_col_ids["Entity"] = self.get_column_by_all_means(sheet=self.ms, key='Entity', default_cols=self.ms_default_cols, label='Entity')
            self.ms_col_ids["Severity"] = self.get_column_by_all_means(sheet=self.ms, key='Severity', default_cols=self.ms_default_cols, label='Severity')
            self.ms_col_ids["CD"] = self.get_column_by_all_means(sheet=self.ms, key='CD', default_cols=self.ms_default_cols, label='Close date')
        
        else:
            cprint('Identifying columns.... (Scansheet ' + self.scan_index + ')')

        self.ss_col_ids["Host"] = self.get_column_by_all_means(sheet=self.ss, key='Host', default_cols=self.ss_default_cols, label='Host', sheet_name='Scansheet ' + self.scan_index)
        self.ss_col_ids["Plugin"] = self.get_column_by_all_means(sheet=self.ss, key='Plugin', default_cols=self.ss_default_cols, label='Plugin', sheet_name='Scansheet ' + self.scan_index)
        self.ss_col_ids["Severity"] = self.get_column_by_all_means(sheet=self.ss, key='Severity', default_cols=self.ss_default_cols, label='Severity', sheet_name='Scansheet ' + self.scan_index)

        # Non mandatory columns

        for key in self.nm_column_keys:
            self.non_mandatory_columns["ss"][key] = self.get_column_by_all_means(sheet=self.ss, key=key, label=column_names[key], sheet_name='Scansheet ' + self.scan_index, default_cols=self.ss_default_cols, important=False)
            self.non_mandatory_columns["ms"][key] = self.get_column_by_all_means(sheet=self.ms, key=key, label=column_names[key], default_cols=self.ms_default_cols, important=False)


    def check_mastersheet_with_scansheet(self):

        cf = 'Carried Forward'
        pcd = 'Patched'

        for ms_row in self.get_column_data(sheet=self.ms, column=self.ms_col_ids["Plugin"]):

            ms_row_str = str(ms_row.row)

            ms_cd_cell = self.ms[self.ms_col_ids["CD"] + ms_row_str]
            ms_status_cell = self.ms[self.ms_col_ids["Status"] + ms_row_str]

            ms_plugin_value = self.trim(ms_row.value).lower()
            ms_host_value = self.trim(self.ms[self.ms_col_ids["Host"] + ms_row_str].value).lower()
            ms_severity_value = self.trim(self.ms[self.ms_col_ids["Severity"] + ms_row_str].value).lower()
            ms_status_value = self.trim(ms_status_cell.value).lower()

            if not (ms_plugin_value or ms_host_value or ms_severity_value): continue

            closed = ms_cd_cell.value or ms_status_value == pcd.lower()

            target_vp = self.vulnerability_param == self.trim(self.ms[self.ms_col_ids["VP"] + ms_row_str].value)
            target_entity = self.entity == self.trim(self.ms[self.ms_col_ids["Entity"] + ms_row_str].value)
            date_created = self.trim(self.ms[self.ms_col_ids["Date"] + ms_row_str].value)

            if not (target_entity and target_vp) or closed: continue

            self.ms_existing_vulnerability_rows.append(ms_row_str)

            if same_week(date_created, self.scan_date): continue

            vulnerability_match = False

            for ss_row in self.get_column_data(sheet=self.ss, column=self.ss_col_ids["Plugin"]):

                if vulnerability_match: continue

                ss_row_str = str(ss_row.row)

                ss_plugin_value = self.trim(ss_row.value).lower()
                ss_host_value = self.trim(self.ss[self.ss_col_ids["Host"] + ss_row_str].value).lower()
                ss_severity_value = self.trim(self.ss[self.ss_col_ids["Severity"] + ss_row_str].value).lower()

                if not (ss_plugin_value or ss_host_value or ss_severity_value): continue

                same_plugin = ms_plugin_value == ss_plugin_value
                same_host = ms_host_value == ss_host_value
                same_severity = ms_severity_value == ss_severity_value

                vulnerability_match = same_host and same_plugin and same_severity

                if (vulnerability_match):

                    ms_ncf_cell = self.ms[self.ms_col_ids["NCF"] + ms_row_str]

                    carried_forward = self.trim(ms_ncf_cell.value).lower() == cf.lower()

                    if not carried_forward:
                        self.set(ms_ncf_cell, cf)
                        self.total_updates["Carried Forward"] += 1
                    break
                    

            if not vulnerability_match:
                self.set(ms_cd_cell, self.scan_date)
                self.total_updates["Closed"] += 1

                if pcd != self.trim(ms_status_cell.value).lower():
                    self.set(ms_status_cell, pcd)


    def check_scansheet_with_mastersheet(self):
        
        for ss_row in self.get_column_data(sheet=self.ss, column=self.ss_col_ids["Plugin"]):

            ss_row_str = str(ss_row.row)
            ss_plugin_value = self.trim(ss_row.value).lower()
            ss_host_value = self.trim(self.ss[self.ss_col_ids["Host"] + ss_row_str].value).lower()
            ss_severity_value = self.trim(self.ss[self.ss_col_ids["Severity"] + ss_row_str].value).lower()

            if not (ss_plugin_value or ss_host_value or ss_severity_value): continue

            vulnerabily_exists = False

            for ms_row in self.ms_existing_vulnerability_rows:
                ms_plugin_value = self.trim(self.ms[self.ms_col_ids["Plugin"] + ms_row].value).lower()
                ms_host_value = self.trim(self.ms[self.ms_col_ids["Host"] + ms_row].value).lower()

                same_plugin = ms_plugin_value == ss_plugin_value 
                same_host = ms_host_value == ss_host_value
                same_severity = self.trim(self.ms[self.ms_col_ids["Severity"] + ms_row].value).lower() == ss_severity_value

                if same_plugin and same_host and same_severity:
                    vulnerabily_exists = True
                    break

            if not vulnerabily_exists:

                ms_last_empty_row = str(len(self.ms['A']) + 1)

                self.set(self.ms[self.ms_col_ids["VP"] + ms_last_empty_row], self.vulnerability_param)
                self.set(self.ms[self.ms_col_ids["Status"] + ms_last_empty_row], 'pending')
                self.set(self.ms[self.ms_col_ids["Date"] + ms_last_empty_row], self.scan_date)
                self.set(self.ms[self.ms_col_ids["Entity"] + ms_last_empty_row], self.entity)
                self.set(self.ms[self.ms_col_ids["NCF"] + ms_last_empty_row], 'New')

                for n in self.nm_column_keys:
                    ss_column = self.non_mandatory_columns['ss'][n]
                    ms_column = self.non_mandatory_columns['ms'][n]
                    if not (ss_column and ms_column):
                        continue
                    self.set(self.ms[self.non_mandatory_columns['ms'][n] + ms_last_empty_row], self.ss[self.non_mandatory_columns['ss'][n] + ss_row_str].value)

                self.total_updates["New"] += 1

    def scan(self):

        ss_path = to_excel(path=self.ss_path)

        cprint('Loading Scansheet ' + self.scan_index + '.....')
        workbook_ss = ex.load_workbook(ss_path)
        cprint('Done!', 'success')

        self.ss = workbook_ss.active
        self.ms = self.workbook_ms.active

        if self.ms_target_sheet:
            try:
                self.ms = self.workbook_ms[self.ms_target_sheet]
            except Exception as e: raise Exception(str(e) + ' (Mastersheet)')
        if self.ss_target_sheet:
            try:
                self.ss = workbook_ss[self.ss_target_sheet]
            except Exception as e: raise Exception(str(e) + ' (Scansheet ' + self.scan_index + ')')

        # Identify columns
        self.get_columns()

        cprint('Done!', 'success')

        # # Check mastersheet with scansheet
        cprint('Scanning & updating Mastersheet with scansheet ' + self.scan_index + '.....')
        self.check_mastersheet_with_scansheet()
        cprint('Done!', 'success')

        # Check scansheet with mastersheet for new vulnerabilities
        cprint('Scanning for new vulnerabilities with scansheet ' + self.scan_index + '.....')
        self.check_scansheet_with_mastersheet()
        cprint('Done!', 'success')

    def save(self, path: str):

        cprint('Saving to ' + path + ' ....')

        if(os.path.exists(path)):
            os.unlink(path)
        self.workbook_ms.save(path)

        # Sometimes, an additional invalid file with no extension is created. So...
        try:
            os.unlink(path.replace('.xlsx', ''))
        except:
            pass
        cprint('Done!', 'success')